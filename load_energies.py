import unicodedata
import re
import openpyxl
import requests

API_BASE = "http://localhost:5001"
ENERGY_ENDPOINT = f"{API_BASE}/api/admin/energies"
WORKBOOK_PATH = "docs/Comparativa combustibles1.xlsb.xlsx"

sheet_to_energy = {
    "GASOIL": "DIESEL",
    "GAS NATURAL": "GAS NATURAL",
    "H2": "H2",
    "HVO": "HVO",
    "BIOMETANO": "BIOMETANO",
    "ELECTRICO": "ELECTRICO",
    "DUOTRAILER GASOIL": "DUO GASOIL",
    "DUOTRAILER HVO": "DUO HVO",
    "DUOTRAILER H2": "DUO H2",
    "DUOTRAILER ELECTRICO": "DUO ELECTRICO",
    "DUOTRAILER BIOMETANO": "DUO BIOMETANO"
}

energy_configs = {
    "DIESEL": {"family": "Diesel", "mode": "Simple", "emission_factor": 2.493},
    "GAS NATURAL": {"family": "GasNatural", "mode": "Simple", "emission_factor": 2.721},
    "H2": {"family": "Hidrogeno", "mode": "Simple", "emission_factor": 0.0},
    "BIOMETANO": {"family": "Biometano", "mode": "Simple", "emission_factor": 2.721, "renewable_share": 1.0, "emission_reduction": 0.9, "emission_reference": "GAS NATURAL"},
    "ELECTRICO": {"family": "Electrico", "mode": "Simple", "emission_factor": 0.0},
    "HVO": {"family": "Hvo", "mode": "Simple", "emission_factor": 2.493, "renewable_share": 1.0, "emission_reduction": 0.9, "emission_reference": "DIESEL"},
    "DUO GASOIL": {"family": "Diesel", "mode": "Duo", "emission_factor": 2.493, "base": "DIESEL", "inherit_emission": True},
    "DUO HVO": {"family": "Hvo", "mode": "Duo", "emission_factor": 2.493, "renewable_share": 1.0, "emission_reduction": 0.9, "base": "HVO", "inherit_emission": True},
    "DUO BIOMETANO": {"family": "Biometano", "mode": "Duo", "emission_factor": 2.721, "renewable_share": 1.0, "emission_reduction": 0.9, "base": "BIOMETANO", "inherit_emission": True},
    "DUO H2": {"family": "Hidrogeno", "mode": "Duo", "emission_factor": 0.0, "base": "H2"},
    "DUO ELECTRICO": {"family": "Electrico", "mode": "Duo", "emission_factor": 0.0, "base": "ELECTRICO"}
}

energy_order = [
    "DIESEL",
    "GAS NATURAL",
    "H2",
    "HVO",
    "BIOMETANO",
    "ELECTRICO",
    "DUO GASOIL",
    "DUO HVO",
    "DUO BIOMETANO",
    "DUO H2",
    "DUO ELECTRICO"
]

def normalize_key(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_only = ''.join(ch for ch in normalized if not unicodedata.combining(ch))
    ascii_only = ascii_only.upper().strip()
    ascii_only = re.sub(r"\s+", " ", ascii_only)
    return ascii_only

def generate_code(name: str) -> str:
    normalized = unicodedata.normalize("NFKD", name)
    ascii_only = ''.join(ch for ch in normalized if not unicodedata.combining(ch))
    ascii_only = ascii_only.upper()
    ascii_only = re.sub(r"[^A-Z0-9]+", "_", ascii_only)
    return ascii_only.strip('_')

def parse_main_table(wb):
    mapping = {}
    ws = wb['CALCULADORA COSTES']
    for row in ws.iter_rows(min_row=2, max_row=20, min_col=1, max_col=6, values_only=True):
        name = row[0]
        if not name:
            continue
        key = normalize_key(str(name))
        if key not in energy_configs:
            continue
        price = float(row[1] or 0)
        consumption = float(row[2] or 0)
        rent = float(row[4] or 0)
        mapping[key] = {
            "price_per_unit": price,
            "consumption_per_100km": consumption * 100,
            "rent": rent
        }
    return mapping

def parse_components(ws):
    monthly_km = None
    components = []
    current_category = None
    order = 0

    for values in ws.iter_rows(min_row=1, max_row=120, min_col=1, max_col=8, values_only=True):
        desc_raw = values[2]
        desc = str(desc_raw).strip() if desc_raw else None
        header = normalize_key(desc) if desc else None

        if header == 'KMS MENSUALES':
            monthly_km = float(values[4] or 0)
            continue
        if header == 'GASTOS FIJOS':
            current_category = 'Fixed'
            continue
        if header == 'GASTOS VARIABLES':
            current_category = 'Variable'
            continue
        if header and header.startswith('COSTE TOTAL'):
            current_category = None
            continue
        if header and header.startswith('PRECIO'):
            continue

        if header == 'PLANIFICACION FLOTA':
            monthly = float(values[4] or 0)
            if monthly:
                components.append({
                    "name": desc,
                    "category": 'Overhead',
                    "valueType": 'MonthlyAmount',
                    "value": round(monthly, 6),
                    "order": order,
                    "isEditable": True
                })
                order += 1
            continue

        if header and header.startswith('MARGEN'):
            percent = values[3]
            if percent:
                components.append({
                    "name": desc,
                    "category": 'Overhead',
                    "valueType": 'PercentageOverSubtotal',
                    "value": round(float(percent), 6),
                    "order": order,
                    "isEditable": True
                })
                order += 1
            continue

        if header in {'ESTRUCTURA INDIRECTA', 'ESTRUCTURA GENERAL'}:
            monthly = float(values[4] or 0)
            if monthly:
                components.append({
                    "name": desc,
                    "category": 'Overhead',
                    "valueType": 'MonthlyAmount',
                    "value": round(monthly, 6),
                    "order": order,
                    "isEditable": True
                })
                order += 1
            continue

        if current_category is None or not desc:
            continue

        monthly_value = float(values[4] or 0)
        rate_value = values[3]
        rate = float(rate_value or 0)

        value_type = 'MonthlyAmount'
        value = monthly_value
        if monthly_km and rate_value is not None:
            if abs(rate * monthly_km - monthly_value) <= 0.5:
                value_type = 'PerKilometerRate'
                value = rate

        if value_type == 'MonthlyAmount' and abs(value) < 1e-6:
            continue
        if value_type == 'PerKilometerRate' and abs(value) < 1e-6:
            continue

        components.append({
            "name": desc,
            "category": current_category,
            "valueType": value_type,
            "value": round(value, 6),
            "order": order,
            "isEditable": True
        })
        order += 1

    return components

def load_components_by_energy(wb):
    mapping = {}
    for ws in wb.worksheets:
        key = normalize_key(ws.title)
        if key not in sheet_to_energy:
            continue
        energy_name = sheet_to_energy[key]
        mapping[energy_name] = parse_components(ws)
    return mapping

def ensure_api_available():
    response = requests.get(ENERGY_ENDPOINT, timeout=5)
    response.raise_for_status()
    return response.json()


def main():
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    base_data = parse_main_table(wb)
    component_data = load_components_by_energy(wb)

    ensure_api_available()

    created = {}

    for energy_name in energy_order:
        config = energy_configs[energy_name]
        base_info = base_data.get(energy_name)
        if base_info is None:
            raise RuntimeError(f"Missing base info for {energy_name}")
        components = component_data.get(energy_name, [])

        payload = {
            "code": generate_code(energy_name),
            "name": energy_name.title(),
            "mode": config['mode'],
            "family": config['family'],
            "pricePerUnit": round(base_info['price_per_unit'], 6),
            "consumptionPer100Km": round(base_info['consumption_per_100km'], 6),
            "rentingCostPerMonth": round(base_info['rent'], 6),
            "emissionFactorPerUnit": config.get('emission_factor', 0.0),
            "renewableShare": config.get('renewable_share'),
            "emissionReduction": config.get('emission_reduction'),
            "baseEnergyId": None,
            "emissionReferenceEnergyId": None,
            "inheritEmissionFromBase": bool(config.get('inherit_emission', False)),
            "isActive": True,
            "costComponents": components
        }

        base_ref = config.get('base')
        if base_ref:
            payload['baseEnergyId'] = created[base_ref]['id']
        emission_ref = config.get('emission_reference')
        if emission_ref:
            payload['emissionReferenceEnergyId'] = created[emission_ref]['id']

        response = requests.post(ENERGY_ENDPOINT, json=payload)
        if response.status_code >= 400:
            print(f"Failed to create {energy_name}: {response.status_code} {response.text}")
        response.raise_for_status()
        data = response.json()
        created[energy_name] = data
        print(f"Created {energy_name}: {data['code']}")

    print("Completed energy import.")

if __name__ == '__main__':
    main()
